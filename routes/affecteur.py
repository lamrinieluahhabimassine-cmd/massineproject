"""
Affecteur routes: bulk file assignment via Excel upload
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, send_file
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from werkzeug.utils import secure_filename
from models import db, User, File, Notification
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io

affecteur_bp = Blueprint('affecteur', __name__, url_prefix='/affecteur')

def affecteur_required(f):
    """Decorator to require affecteur role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            abort(401)
        if current_user.role not in ['affecteur', 'admin']:
            flash('Accès réservé aux affecteurs.', 'danger')
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

@affecteur_bp.before_request
@login_required
@affecteur_required
def require_affecteur():
    """All affecteur routes require affecteur role"""
    pass


@affecteur_bp.route('/dashboard')
def dashboard():
    """Affecteur dashboard"""
    # Get upload history (files created by this affecteur)
    recent_uploads = File.query.order_by(File.created_at.desc()).limit(50).all()
    
    # Statistics
    stats = {
        'total_assigned': File.query.count(),
        'assigned_today': File.query.filter(
            File.created_at >= datetime.now().replace(hour=0, minute=0, second=0)
        ).count(),
        'pending_evaluation': File.query.filter_by(status='en attente d\'évaluation').count(),
        'total_users': User.query.filter_by(role='user').count(),
    }
    
    return render_template('affecteur/dashboard.html',
                         recent_uploads=recent_uploads,
                         stats=stats)


@affecteur_bp.route('/upload', methods=['GET', 'POST'])
def upload_excel():
    """Upload Excel file for bulk assignment"""
    if request.method == 'POST':
        # Check if file was uploaded
        if 'excel_file' not in request.files:
            flash('Aucun fichier sélectionné.', 'danger')
            return render_template('affecteur/upload_form.html')
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('Aucun fichier sélectionné.', 'danger')
            return render_template('affecteur/upload_form.html')
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Format de fichier invalide. Utilisez .xlsx ou .xls', 'danger')
            return render_template('affecteur/upload_form.html')
        
        try:
            # Load Excel file using openpyxl
            from openpyxl import load_workbook
            workbook = load_workbook(file)
            worksheet = workbook.active
            
            # Get header row (first row)
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value.lower().strip() if cell.value else None)
            
            # Validate required columns (WITHOUT user_email)
            required_columns = ['file_number', 'receipt_date', 'importer', 'exporter', 
                              'country', 'route']
            
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                flash(f'Colonnes manquantes dans le fichier Excel: {", ".join(missing_columns)}', 'danger')
                return render_template('affecteur/upload_form.html')
            
            # Get column indices
            col_indices = {header: idx + 1 for idx, header in enumerate(headers) if header}
            
            # Process each row (starting from row 2)
            success_count = 0
            error_count = 0
            errors = []
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract values from row
                    file_number = str(row[col_indices['file_number'] - 1]).strip() if col_indices.get('file_number') else None
                    receipt_date = row[col_indices['receipt_date'] - 1] if col_indices.get('receipt_date') else None
                    importer = str(row[col_indices['importer'] - 1]).strip() if col_indices.get('importer') else None
                    exporter = str(row[col_indices['exporter'] - 1]).strip() if col_indices.get('exporter') else None
                    country = str(row[col_indices['country'] - 1]).strip() if col_indices.get('country') else None
                    route = str(row[col_indices['route'] - 1]).strip().upper() if col_indices.get('route') else None
                    sor_number = str(row[col_indices['sor_number'] - 1]).strip() if col_indices.get('sor_number') and row[col_indices['sor_number'] - 1] else None
                    sol_number = str(row[col_indices['sol_number'] - 1]).strip() if col_indices.get('sol_number') and row[col_indices['sol_number'] - 1] else None
                    
                    # Skip empty rows
                    if not file_number or file_number == 'None':
                        continue
                    
                    # Check if file number already exists
                    if File.query.filter_by(file_number=file_number).first():
                        errors.append(f"Ligne {row_idx}: Numéro de dossier {file_number} existe déjà")
                        error_count += 1
                        continue
                    
                    # Parse date
                    if isinstance(receipt_date, str):
                        receipt_date = datetime.strptime(receipt_date, '%Y-%m-%d').date()
                    elif hasattr(receipt_date, 'date'):
                        receipt_date = receipt_date.date()
                    else:
                        receipt_date = receipt_date
                    
                    # Validate route requirements
                    if route == 'B' and not sor_number:
                        errors.append(f"Ligne {row_idx}: SOR requis pour Route B")
                        error_count += 1
                        continue
                    
                    if route == 'C' and not sol_number:
                        errors.append(f"Ligne {row_idx}: SOL requis pour Route C")
                        error_count += 1
                        continue
                    
                    # Create file (user_id is NULL - not assigned)
                    new_file = File(
                        file_number=file_number,
                        receipt_date=receipt_date,
                        importer=importer,
                        exporter=exporter,
                        country=country,
                        route=route,
                        sor_number=sor_number,
                        sol_number=sol_number,
                        status='en attente d\'évaluation',
                        user_id=None  # Not assigned to anyone
                    )
                    
                    db.session.add(new_file)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Ligne {row_idx}: Erreur - {str(e)}")
                    error_count += 1
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Show results
            if success_count > 0:
                flash(f'✅ {success_count} dossier(s) créé(s) avec succès!', 'success')
            
            if error_count > 0:
                flash(f'⚠️ {error_count} erreur(s) détectée(s).', 'warning')
                for error in errors[:10]:  # Show first 10 errors
                    flash(error, 'danger')
                if len(errors) > 10:
                    flash(f'... et {len(errors) - 10} autres erreurs.', 'danger')
            
            if success_count > 0:
                return redirect(url_for('affecteur.dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erreur lors du traitement du fichier: {str(e)}', 'danger')
            return render_template('affecteur/upload_form.html')
    
    return render_template('affecteur/upload_form.html')


@affecteur_bp.route('/template/download')
def download_template():
    """Download Excel template"""
    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Dossiers'
    
    # Add headers (WITHOUT user_email)
    headers = ['file_number', 'receipt_date', 'importer', 'exporter', 
               'country', 'route', 'sor_number', 'sol_number']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
    
    # Add sample data (WITHOUT user_email)
    sample_data = [
        ['VOC-2025-001', '2025-11-09', 'Company A', 'Exporter X', 'France', 'A', '', ''],
        ['VOC-2025-002', '2025-11-09', 'Company B', 'Exporter Y', 'Spain', 'B', 'SOR-123', ''],
        ['VOC-2025-003', '2025-11-09', 'Company C', 'Exporter Z', 'Italy', 'C', '', 'SOL-456'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='template_affectation_dossiers.xlsx'
    )


@affecteur_bp.route('/history')
def history():
    """View upload history"""
    files = File.query.order_by(File.created_at.desc()).all()
    
    return render_template('affecteur/history.html', files=files)